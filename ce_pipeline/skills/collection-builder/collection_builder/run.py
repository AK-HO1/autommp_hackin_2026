#!/usr/bin/env python3
"""
run.py — entry point for the collection builder.

Usage:
    export OPENAI_API_KEY=sk-...
    python run.py --feed feed.csv --ce-list existing_ces.csv \
                  --out decisions.csv [--xlsx decisions.xlsx] [--limit N]

--feed      CSV of incoming experiences. Columns (configurable below):
              experience_name   (required)
              link              (optional)
              experience_id     (optional; derived from link/index if absent)
--ce-list   CSV of existing collections to match against. Columns:
              name (or "Combined Entity Name")
              city (or parsed from "<...> - City")
--out       output CSV (the labelling template: has correct_collection + notes)
--xlsx      optional formatted spreadsheet
--limit     process only the first N feed rows (for a quick test)

FAILS LOUD: no API key or a failed API call aborts the run. There is no
TF-IDF/offline fallback — the numbers you get are always from the real model.
"""
import argparse
import csv
import os
import re
import sys

# allow running as `python run.py` from the package's parent dir
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from collection_builder import engine
from collection_builder.engine import CollectionMemory

FEED_NAME_COLS = ["experience_name", "Product Name", "name", "title"]
FEED_LINK_COLS = ["link", "Link", "url", "URL"]
FEED_ID_COLS = ["experience_id", "id", "product_id"]
CE_NAME_COLS = ["name", "Combined Entity Name", "collection_name", "Collection Name"]
CE_CITY_COLS = ["city", "City", "Primary City", "primary_city"]
CE_ID_COLS = ["Combined Entity ID", "id", "collection_id", "ID"]


def _pick(row, candidates):
    for c in candidates:
        if c in row and str(row[c]).strip():
            return str(row[c]).strip()
    return ""


def load_feed(path, limit=None):
    rows = list(csv.DictReader(open(path, encoding="utf-8-sig")))
    if limit:
        rows = rows[:limit]
    out = []
    for i, r in enumerate(rows, 1):
        name = _pick(r, FEED_NAME_COLS)
        if not name:
            continue
        link = _pick(r, FEED_LINK_COLS)
        eid = _pick(r, FEED_ID_COLS)
        if not eid:
            m = re.search(r"-t(\d+)/", link)
            eid = m.group(1) if m else f"row{i}"
        out.append({"experience_id": eid, "experience_name": name, "link": link})
    return out


def _city_from_ce_name(name):
    # "Guided Tours - Auckland" -> Auckland ; "Sky Tower" -> "" (POI, city unknown)
    if " - " in name:
        return name.split(" - ")[-1].strip()
    return ""


def load_ce_list(path):
    rows = list(csv.DictReader(open(path, encoding="utf-8-sig")))
    entries = []
    for r in rows:
        name = _pick(r, CE_NAME_COLS)
        if not name:
            continue
        city = _pick(r, CE_CITY_COLS) or _city_from_ce_name(name)
        cid = _pick(r, CE_ID_COLS)
        entries.append({"name": name, "city": city, "id": cid})
    return entries


OUT_COLS = ["experience_id", "experience_name", "primary_city", "theme",
            "decision", "collection_id", "collection_name", "is_new",
            "confidence", "needs_review", "reason", "link",
            "correct_collection", "notes"]


def write_csv(records, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=OUT_COLS)
        w.writeheader()
        for r in records:
            w.writerow({k: r.get(k, "") for k in OUT_COLS})


def write_xlsx(records, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("FATAL: openpyxl not installed. Run: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "Decisions"
    ws.append(OUT_COLS)
    navy, green, amber = "1F3864", "C6EFCE", "FFEB9C"
    for c in range(1, len(OUT_COLS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=navy)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in records:
        ws.append([r.get(k, "") for k in OUT_COLS])
        row = ws.max_row
        if r.get("needs_review"):
            ws.cell(row, OUT_COLS.index("needs_review") + 1).fill = \
                PatternFill("solid", start_color=amber)
        if r.get("decision") == "assign_existing":
            ws.cell(row, OUT_COLS.index("decision") + 1).fill = \
                PatternFill("solid", start_color=green)
    widths = [12, 46, 16, 18, 18, 22, 30, 8, 10, 12, 40, 30, 26, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


def summarize(records):
    from collections import Counter
    dec = Counter(r["decision"] for r in records)
    rev = sum(1 for r in records if r["needs_review"])
    print("\n=== SUMMARY ===", file=sys.stderr)
    print(f"  total: {len(records)}", file=sys.stderr)
    for k, v in dec.most_common():
        print(f"  {k}: {v}", file=sys.stderr)
    print(f"  needs_review: {rev}", file=sys.stderr)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--feed", required=True)
    ap.add_argument("--ce-list", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--xlsx")
    ap.add_argument("--limit", type=int)
    ap.add_argument("--market", default="",
                    help="Comma-separated cities/region to scope the CE list to "
                         "(e.g. 'New Zealand' cities). If omitted, the CE list is "
                         "auto-scoped to the cities the gate detects in the feed, "
                         "with no scope only as a last resort.")
    args = ap.parse_args()

    feed = load_feed(args.feed, args.limit)
    ce = load_ce_list(args.ce_list)
    if not feed:
        sys.exit("FATAL: no experiences loaded from --feed")
    print(f"loaded {len(feed)} experiences, {len(ce)} existing collections",
          file=sys.stderr)

    # Market scoping: explicit --market wins. Otherwise we DON'T pre-scope here
    # (the gate hasn't run yet to know cities); the same-city + same-market
    # fallback in Job A handles it. Pre-scoping is mainly a speed/safety win for
    # huge global lists — pass --market for that.
    market_cities = None
    if args.market:
        market_cities = [c.strip() for c in args.market.split(",") if c.strip()]
        print(f"scoping CE list to market cities: {market_cities}", file=sys.stderr)

    memory = CollectionMemory(ce, market_cities=market_cities)
    print(f"  -> {len(memory.entries)} named POIs + {len(memory.buckets)} buckets "
          f"in scope", file=sys.stderr)
    records = engine.run(feed, memory)

    write_csv(records, args.out)
    print(f"wrote {args.out}", file=sys.stderr)
    if args.xlsx:
        write_xlsx(records, args.xlsx)
        print(f"wrote {args.xlsx}", file=sys.stderr)
    summarize(records)


if __name__ == "__main__":
    main()
