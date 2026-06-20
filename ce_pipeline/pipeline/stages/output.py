"""Output stage. Renders CEAssortment objects into:

  1. experienceos_upload.csv  -> the flat, machine-uploadable file for ExperienceOS
  2. assortment_review.xlsx   -> a 3-sheet workbook mirroring the manual format
                                 (Assortment / Cross-sell Bench / Logic & Notes)

The CSV is FLAT (every row self-contained) rather than visually merged like the
manual sheet, because a machine upload needs each row to carry its full key.
A `ce_name` column is added up front so one file can hold an entire city's CEs.
"""
from __future__ import annotations
import csv
from typing import Iterable
from ..schema import CEAssortment, Experience, Variant

CSV_COLUMNS = [
    "ce_name", "ce_id", "is_new_ce",
    "experience_rank", "experience_name",
    "variant_rank", "variant_name", "variant_content",
    "product_name", "supplier_name",
    "confidence", "is_new_variant", "comments",
]


def _supplier_cell(v: Variant) -> str:
    """Render suppliers the way the manual sheet does:
    'GlobalTix (290941, 290947) / BeMyGuest (59324)'."""
    parts = []
    for s in v.suppliers:
        ids = ", ".join(s.product_ids)
        tag = f" [{s.note}]" if s.note else ""
        parts.append(f"{s.sp_name} ({ids}){tag}" if ids else f"{s.sp_name}{tag}")
    return " / ".join(parts)


def assortment_to_rows(a: CEAssortment) -> list[dict]:
    rows = []
    for exp in a.experiences:
        for var in exp.variants:
            rows.append({
                "ce_name": a.ce_name,
                "ce_id": a.ce_id or "",
                "is_new_ce": "Y" if a.is_new_ce else "",
                "experience_rank": exp.rank,
                "experience_name": exp.name,
                "variant_rank": var.rank,
                "variant_name": var.name,
                "variant_content": var.content,
                "product_name": var.product_name,
                "supplier_name": _supplier_cell(var),
                "confidence": "" if var.confidence is None else round(var.confidence, 2),
                "is_new_variant": "Y" if var.is_new else "",
                "comments": var.comments,
            })
    return rows


def write_csv(assortments: Iterable[CEAssortment], path: str) -> int:
    rows = [r for a in assortments for r in assortment_to_rows(a)]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        w.writeheader()
        w.writerows(rows)
    return len(rows)


def write_review_xlsx(assortments: list[CEAssortment], path: str) -> None:
    """Optional reviewer workbook that mirrors the manual 3-sheet layout
    (visually merged Experience cells). Requires openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Assortment"
    hdr = ["CE", "Experience Rank", "Experience Name", "Variant Rank",
           "Variant Name", "Variant Content", "Product Name",
           "Supplier Name", "Confidence", "Comments"]
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True)
    gold = PatternFill("solid", start_color="FFF2CC")
    for a in assortments:
        for exp in a.experiences:
            first = True
            for var in exp.variants:
                ws.append([
                    a.ce_name if first else "",
                    exp.rank if first else "",
                    exp.name if first else "",
                    var.rank, var.name, var.content, var.product_name,
                    _supplier_cell(var),
                    "" if var.confidence is None else round(var.confidence, 2),
                    var.comments,
                ])
                if var.is_new:
                    for c in ws[ws.max_row]:
                        c.fill = gold
                first = False

    bench = wb.create_sheet("Cross-sell Bench")
    bench.append(["CE", "Product (available supply)", "Indicative price",
                  "Where it would sit", "Why bench it"])
    for c in bench[1]:
        c.font = Font(bold=True)
    for a in assortments:
        for b in a.bench:
            bench.append([a.ce_name, b.product, b.indicative_price,
                          b.where_it_would_sit, b.why_bench])

    notes = wb.create_sheet("Logic & Notes")
    notes.append(["CE", "Topic", "Detail"])
    for c in notes[1]:
        c.font = Font(bold=True)
    for a in assortments:
        for k, val in a.notes.items():
            notes.append([a.ce_name, k, val])

    wb.save(path)
